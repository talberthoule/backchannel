import uuid
from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from unittest import IsolatedAsyncioTestCase
from unittest.mock import AsyncMock, patch

from openpyxl import load_workbook

from app.routers import artifacts
from app.routers.artifacts import export_insights, export_summary


class _Result:
    def __init__(self, values):
        self._values = values

    def scalars(self):
        return self

    def all(self):
        return self._values


class _Db:
    def __init__(self, questions, speakers=(), session=None):
        self.session = session or SimpleNamespace(
            name="Client Review", started_at=None, ended_at=None
        )
        self._results = iter([_Result(questions), _Result(list(speakers))])
        self.expunged = []

    async def get(self, _model, _session_id):
        return self.session

    async def execute(self, _query):
        return next(self._results)

    def expunge(self, row):
        self.expunged.append(row)


def _speaker(name, *, role="", speaker_type="external", display_name="", display_enabled=False):
    return SimpleNamespace(
        id=uuid.uuid4(),
        name=name,
        role=role,
        speaker_type=speaker_type,
        display_name=display_name,
        display_name_enabled=display_enabled,
        created_at=datetime(2026, 8, 1, tzinfo=timezone.utc),
    )


def _question(text: str, *, enhanced: bool = False, item_type="observation", **overrides):
    row = SimpleNamespace(
        item_type=item_type,
        lens_label="Discovery",
        question=text,
        rationale="Why it matters",
        source_context="It was said out loud",
        speaker_id=None,
        vote=0,
        starred=False,
        answered=False,
        answer_summary="",
        needs_followup=False,
        followup_question="",
        offering_match="",
        enhanced=enhanced,
        dismissed=False,
        enrichment_notes="",
        agent_source="consolidated_analyst",
        revision_count=1 if enhanced else 0,
        created_at=datetime(2026, 8, 1, tzinfo=timezone.utc),
    )
    for key, value in overrides.items():
        setattr(row, key, value)
    return row


async def _workbook(db, **kwargs):
    response = await export_insights(uuid.uuid4(), db=db, **kwargs)
    content = b"".join([chunk async for chunk in response.body_iterator])
    return load_workbook(BytesIO(content)), response


def _cells(ws):
    return [value for row in ws.iter_rows(values_only=True) for value in row if value is not None]


def _joined(ws):
    return " ".join(str(value) for value in _cells(ws))


class InsightsExportTests(IsolatedAsyncioTestCase):
    async def test_export_contains_each_base_and_enhanced_insight_once(self):
        db = _Db([
            _question("Base insight", enhanced=False),
            _question("Enhanced insight", enhanced=True),
        ])

        wb, response = await _workbook(db)
        rows = list(wb["Insights"].iter_rows(values_only=True))
        headers = rows[0]
        insight_col = headers.index("Insight")
        enhanced_col = headers.index("Enhanced")

        self.assertEqual(
            'attachment; filename="insights-Client_Review.xlsx"',
            response.headers["content-disposition"],
        )
        self.assertEqual(2, len(rows) - 1)
        self.assertEqual(
            {("Base insight", False), ("Enhanced insight", True)},
            {(row[insight_col], row[enhanced_col]) for row in rows[1:]},
        )

    async def test_the_workbook_opens_on_a_summary_and_keeps_the_grid_intact(self):
        # ALP-369: the reader-facing sheets are additive. Insights stays the
        # record of eighteen columns; it just is not the first thing seen.
        db = _Db([_question("Base insight")])
        wb, _ = await _workbook(db)

        self.assertEqual(["Summary", "Action Items", "Opportunities", "Insights"], wb.sheetnames)
        self.assertEqual("Summary", wb.active.title)
        self.assertEqual(18, len(list(wb["Insights"].iter_rows(values_only=True))[0]))

    async def test_the_summary_reports_the_mix_participants_and_what_is_open(self):
        alice = _speaker("Voice 1", role="CISO", speaker_type="external")
        db = _Db(
            [
                _question("Migrate the estate", item_type="action_item", speaker_id=alice.id),
                _question("Price is high", item_type="objection"),
                _question("What is the timeline?", item_type="question"),
                _question("Answered already", item_type="question", answered=True),
                _question("Starred thing", starred=True),
            ],
            speakers=[alice],
        )
        wb, _ = await _workbook(db)
        cells = _cells(wb["Summary"])

        self.assertIn("BACKCHANNEL BRIEFING", cells)
        self.assertIn("Client Review", cells)
        self.assertIn("Action Item", cells)
        self.assertIn("Objection", cells)
        self.assertIn("Voice 1", cells)
        self.assertIn("CISO", cells)
        self.assertIn("External", cells)
        # Starred, then the objection, then the open question. An answered
        # question is not something that needs attention.
        self.assertIn("Starred", cells)
        self.assertIn("Objection raised", cells)
        self.assertIn("Unanswered", cells)
        self.assertIn("What is the timeline?", cells)
        self.assertNotIn("Answered already", cells)

    async def test_dismissed_insights_stay_in_the_grid_and_out_of_the_summary(self):
        db = _Db([
            _question("Kept", item_type="action_item"),
            _question("Rejected", item_type="action_item", dismissed=True),
        ])
        wb, _ = await _workbook(db)

        grid = _cells(wb["Insights"])
        self.assertIn("Kept", grid)
        self.assertIn("Rejected", grid)

        actions = _cells(wb["Action Items"])
        self.assertIn("Kept", actions)
        self.assertNotIn("Rejected", actions)
        self.assertIn("1 captured", actions)

    async def test_the_opportunities_sheet_pairs_each_one_with_its_match(self):
        db = _Db([
            _question(
                "Refresh the WAN",
                item_type="opportunity",
                offering_match="Cisco SD-WAN, managed",
                enrichment_notes="Matched on the branch count",
            ),
            _question("Unclear need", item_type="opportunity"),
        ])
        wb, _ = await _workbook(db)
        cells = _cells(wb["Opportunities"])

        self.assertIn("Cisco SD-WAN, managed", cells)
        self.assertIn("Matched on the branch count", cells)
        self.assertIn("Not matched", cells)
        self.assertTrue(any("1 matched" in str(value) for value in cells))

    async def test_display_names_are_substituted_on_every_sheet(self):
        renamed = _speaker("Voice 1", role="CISO", display_name="Dana Client", display_enabled=True)
        db = _Db(
            [_question("Voice 1 owns the migration", item_type="action_item", speaker_id=renamed.id)],
            speakers=[renamed],
        )
        wb, _ = await _workbook(db)

        for sheet in ("Summary", "Action Items", "Insights"):
            joined = _joined(wb[sheet])
            self.assertIn("Dana Client", joined, sheet)
            self.assertNotIn("Voice 1", joined, sheet)


class RevealedExportTests(IsolatedAsyncioTestCase):
    """A new worksheet must never emit a vault value the export was not asked
    for, and must show one when it was (ALP-369)."""

    def _db(self):
        speaker = _speaker("[PERSON_2]", role="CISO")
        return _Db(
            [
                _question(
                    "[PERSON_1] will migrate [ORG_1]",
                    item_type="action_item",
                    speaker_id=speaker.id,
                    starred=True,
                )
            ],
            speakers=[speaker],
        )

    async def test_tokens_stay_tokens_without_reveal(self):
        wb, _ = await _workbook(self._db())
        for sheet in ("Summary", "Action Items", "Insights"):
            joined = _joined(wb[sheet])
            self.assertIn("[PERSON_1]", joined, sheet)
            self.assertNotIn("Sarah Connor", joined, sheet)

    async def test_reveal_reaches_the_summary_sheets_and_the_roster(self):
        mapping = {"[PERSON_1]": "Sarah Connor", "[ORG_1]": "Cyberdyne", "[PERSON_2]": "Dana Client"}
        recorded = []

        with (
            patch.object(artifacts.vault, "reveal_map", new=AsyncMock(return_value=mapping)),
            patch.object(
                artifacts.shield,
                "record_reveal",
                new=AsyncMock(side_effect=lambda *args: recorded.append(args)),
            ),
        ):
            wb, _ = await _workbook(self._db(), reveal=True)

        for sheet in ("Summary", "Action Items", "Insights"):
            joined = _joined(wb[sheet])
            self.assertIn("Sarah Connor", joined, sheet)
            self.assertNotIn("[PERSON_1]", joined, sheet)
        # The roster label the summary prints is revealed too, and every
        # substitution is counted in the audit record.
        self.assertIn("Dana Client", _joined(wb["Summary"]))
        self.assertEqual(1, len(recorded))
        self.assertEqual("questions-export", recorded[0][1])
        self.assertGreaterEqual(recorded[0][2], 3)


# ---------------------------------------------------------------------------
# The HTML briefing (ALP-370)
# ---------------------------------------------------------------------------


class _SummaryDb:
    """Loads the briefing's supporting rows, and refuses to serve a synthesis.

    The synthesis must come from get_session_synthesis, which applies the
    speaker-mapping normalization; a direct query would skip it. Everything
    else the document draws is a legitimate query.
    """

    def __init__(self, session, questions=(), speakers=(), transcript=(), directives=()):
        self.session = session
        self._results = iter([
            _Result(list(questions)),
            _Result(list(speakers)),
            _Result(list(transcript)),
            _Result(list(directives)),
        ])
        self.queries = 0

    async def get(self, _model, _session_id):
        return self.session

    async def execute(self, query):
        assert "session_syntheses" not in str(query), "briefing queried the synthesis directly"
        self.queries += 1
        return next(self._results)


def _session(name="Client Review", meeting_type="client_sales"):
    return SimpleNamespace(
        name=name,
        meeting_type=meeting_type,
        started_at=datetime(2026, 9, 2, 14, 2, tzinfo=timezone.utc),
        ended_at=datetime(2026, 9, 2, 14, 59, tzinfo=timezone.utc),
    )


def _turn(text, speaker=None, minute=2):
    return SimpleNamespace(
        text=text,
        timestamp=datetime(2026, 9, 2, 14, minute, tzinfo=timezone.utc),
        speaker_id=speaker.id if speaker else None,
    )


def _synthesis(status="completed", **overrides):
    row = SimpleNamespace(
        status=status,
        top_outcomes=[{
            "title": "The renewal is a replacement",
            "summary": "The MPLS contract ends in March and will not be renewed.",
            "rationale": "Stated twice, unprompted.",
        }],
        client_objectives=[{"title": "Exit MPLS cleanly"}],
        top_opportunities=[
            {"title": "Managed SD-WAN", "summary": "40 branches", "owner": "Marcus", "status": "Qualifying"}
        ],
        risks_blockers=[{"title": "No named decision owner", "summary": "The sponsor was not on the call."}],
        action_plan=[{"title": "Confirm the branch count", "owner": "Marcus", "status": "Open"}],
        unresolved_discovery_questions=[{"title": "Who owns the standard?"}],
        strategic_signals=[],
        clusters=[],
        arbiter_notes="The lenses agreed on the deadline.",
    )
    for key, value in overrides.items():
        setattr(row, key, value)
    return row


async def _html(db, synthesis, **kwargs):
    with patch.object(
        artifacts, "get_session_synthesis", new=AsyncMock(return_value=synthesis)
    ) as getter:
        response = await export_summary(uuid.uuid4(), db=db, **kwargs)
        content = b"".join([chunk async for chunk in response.body_iterator])
    return content.decode("utf-8"), response, getter


class BriefingExportTests(IsolatedAsyncioTestCase):
    async def test_the_synthesized_briefing_leads_with_the_answer(self):
        speaker = _speaker("Voice 1", role="CISO")
        db = _SummaryDb(
            _session(),
            questions=[
                _question(
                    "Confirm the count",
                    item_type="action_item",
                    speaker_id=speaker.id,
                    starred=True,
                )
            ],
            speakers=[speaker],
            transcript=[_turn("The MPLS deal runs out in March.", speaker)],
            directives=[SimpleNamespace(text="Find out who owns the standard.")],
        )
        html, response, getter = await _html(db, _synthesis())

        getter.assert_awaited_once()
        self.assertEqual(
            'attachment; filename="briefing-Client_Review.html"',
            response.headers["content-disposition"],
        )
        # The answer sits above everything: the standfirst is the leading
        # outcome's thesis, and the tally counts what came out of the call.
        self.assertIn('<p class="lede">The renewal is a replacement</p>', html)
        self.assertIn("ACTIONS", html.upper())
        # The section that supplied the standfirst does not repeat it.
        self.assertEqual(1, html.count("The renewal is a replacement"))
        # Sections named for the meeting that happened.
        self.assertIn("Client objectives", html)
        # Evidence is reachable from the insight, through a native disclosure.
        self.assertIn("<details", html)
        self.assertIn("It was said out loud", html)
        # So is the transcript, and the directives that framed the call.
        self.assertIn("Full transcript", html)
        self.assertIn("Find out who owns the standard.", html)
        self.assertIn("Voice 1", html)

    async def test_a_partial_synthesis_says_so_instead_of_reading_as_complete(self):
        db = _SummaryDb(_session())
        html, _, _ = await _html(db, _synthesis(status="partial"))
        self.assertIn("partial synthesis", html)

    async def test_no_synthesis_gets_the_same_document_led_by_the_insights(self):
        # ALP-370: the fallback is not a plainer skin. Same shell, same
        # figures, same evidence rule; only the source of the sections differs.
        speaker = _speaker("Voice 1")
        db = _SummaryDb(
            _session(),
            questions=[
                _question("Confirm the count", item_type="action_item", speaker_id=speaker.id),
                _question("Who signs off?", item_type="question", speaker_id=speaker.id),
            ],
            speakers=[speaker],
            transcript=[_turn("We need to check that.", speaker)],
        )
        html, response, _ = await _html(db, None)

        self.assertEqual(
            'attachment; filename="summary-Client_Review.html"',
            response.headers["content-disposition"],
        )
        self.assertIn("call record", html)
        self.assertIn("Commitments", html)
        self.assertIn("Still open", html)
        self.assertIn("Confirm the count", html)
        self.assertIn("Full transcript", html)
        for marker in ('class="band"', 'class="lede"', 'class="colophon"', 'class="tally"'):
            self.assertIn(marker, html)

    async def test_an_unusable_synthesis_falls_back_rather_than_rendering_empty(self):
        db = _SummaryDb(_session(), questions=[_question("Something")])
        empty = _synthesis(
            status="completed",
            top_outcomes=[],
            top_opportunities=[],
            risks_blockers=[],
            action_plan=[],
            unresolved_discovery_questions=[],
            client_objectives=[],
            clusters=[],
        )
        html, response, _ = await _html(db, empty)
        self.assertIn("call record", html)
        self.assertIn('filename="summary-Client_Review.html"', response.headers["content-disposition"])

    async def test_the_document_is_self_contained_and_prints(self):
        db = _SummaryDb(_session(), questions=[_question("Something")])
        html, _, _ = await _html(db, _synthesis())

        # No network of any kind: the file has to render offline years from now.
        for forbidden in ("http://", "https://", "<script", "<img", "@import", "url("):
            self.assertNotIn(forbidden, html, forbidden)
        self.assertIn("@media print", html)
        self.assertIn("prefers-color-scheme: dark", html)
        self.assertTrue(html.startswith("<!DOCTYPE html>"))
        self.assertTrue(html.rstrip().endswith("</html>"))

    async def test_the_figures_fall_away_when_they_would_say_nothing(self):
        # One insight type is not a mix, and one speaker is not participation.
        solo = _speaker("Voice 1")
        db = _SummaryDb(
            _session(),
            questions=[
                _question("A", item_type="observation"),
                _question("B", item_type="observation"),
            ],
            speakers=[solo],
            transcript=[_turn("Talking.", solo)],
        )
        html, _, _ = await _html(db, None)
        self.assertNotIn("What was captured", html)
        self.assertNotIn("Turns taken", html)

    async def test_html_special_characters_in_stored_text_are_escaped(self):
        db = _SummaryDb(
            _session(name="Q&A <review>"),
            questions=[_question('A "quote" and <b>bold</b>', item_type="action_item")],
        )
        html, _, _ = await _html(db, None)
        self.assertIn("&lt;b&gt;bold&lt;/b&gt;", html)
        self.assertNotIn("<b>bold</b>", html)
        self.assertIn("Q&amp;A &lt;review&gt;", html)


class RevealedBriefingTests(IsolatedAsyncioTestCase):
    """The shield reveals the rendered document as a whole, so every token has
    to survive rendering intact, and none may resolve without the flag."""

    def _db(self):
        speaker = _speaker("[PERSON_2]", role="CISO")
        return _SummaryDb(
            _session(),
            questions=[
                _question(
                    "[PERSON_1] will migrate [ORG_1]",
                    item_type="action_item",
                    speaker_id=speaker.id,
                    starred=True,
                    source_context="[PERSON_1] said the estate moves in March",
                )
            ],
            speakers=[speaker],
            transcript=[_turn("[PERSON_1] said the estate moves in March", speaker)],
        )

    async def test_tokens_stay_tokens_without_reveal(self):
        html, _, _ = await _html(self._db(), _synthesis())
        self.assertIn("[PERSON_1]", html)
        self.assertIn("[PERSON_2]", html)
        self.assertNotIn("Sarah Connor", html)

    async def test_every_token_reaches_the_page_in_a_revealable_form(self):
        async def fake_reveal(_db, _session_id, text, route=""):
            self.assertEqual("summary-export", route)
            for token, value in (
                ("[PERSON_1]", "Sarah Connor"),
                ("[PERSON_2]", "Dana Client"),
                ("[ORG_1]", "Cyberdyne"),
            ):
                self.assertIn(token, text, f"{token} did not survive rendering")
                text = text.replace(token, value)
            return text

        with patch.object(artifacts.shield, "reveal_text", new=AsyncMock(side_effect=fake_reveal)):
            html, _, _ = await _html(self._db(), _synthesis(), reveal=True)

        self.assertIn("Sarah Connor", html)
        self.assertIn("Dana Client", html)
        self.assertNotIn("[PERSON_", html)
