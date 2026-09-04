"""The insights workbook: a summary someone will actually read, then the grid.

``Insights`` is the system of record - eighteen columns, every type in one
table, nothing dropped. It is the right shape for a filter and the wrong shape
for a person who has ninety seconds. The three sheets in front of it are that
person's view: what happened and in what proportion, who was in the room, what
still needs a decision, and then the two lists anyone actually works from.

Nothing here reads the database. It is handed already-loaded rows and an
already-resolved speaker roster, which is what keeps the reveal rule simple:
the router substitutes vault values once, on detached rows, before calling in.
Whatever reaches these sheets is whatever the caller decided the reader may
see, so a new worksheet cannot leak a value the export was not asked for
(ALP-369).
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime

from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# The product palette, straight from frontend/src/index.css primitives, so an
# exported workbook and the app it came from are recognizably the same thing.
INK = "0F172A"          # slate-900
INK_SOFT = "475569"     # slate-600
INK_FAINT = "64748B"    # slate-500
TEAL = "0D9488"
TEAL_DARK = "0F766E"
TEAL_WASH = "E7F4F2"    # a tint of TEAL, for header bands
LINE = "CBD5E1"         # slate-300
CANVAS = "F8FAFC"       # slate-50
AMBER = "B45309"

_TYPE_LABELS = {
    "question": "Question",
    "observation": "Observation",
    "opportunity": "Opportunity",
    "action_item": "Action Item",
    "objection": "Objection",
    "asked": "Asked",
}

# Reading order for the mix table: what was decided, then what was found, then
# what was asked. Types outside this list keep their relative order after it.
_TYPE_ORDER = ["action_item", "opportunity", "objection", "observation", "question", "asked"]

_THIN = Side(style="thin", color=LINE)
_BOTTOM_RULE = Border(bottom=Side(style="thin", color=LINE))
_HEADER_BORDER = Border(bottom=Side(style="medium", color=TEAL))


def type_label(item_type: str | None) -> str:
    slug = item_type or "question"
    return _TYPE_LABELS.get(slug, slug.replace("_", " ").title())


@dataclass(frozen=True)
class Participant:
    """One speaker as the workbook shows them - never an ORM row.

    ``label`` is whatever the caller resolved: a display name, the auto name,
    or a vault token when the export was not asked to reveal.
    """

    id: str
    label: str
    role: str
    side: str


def _write(
    ws: Worksheet,
    row: int,
    values: list,
    *,
    bold: bool = False,
    size: int = 11,
    color: str = INK,
    fill: str | None = None,
    wrap: bool = False,
    border: Border | None = None,
    top: str = "top",
) -> int:
    """Write one row and return the next row index."""
    for offset, value in enumerate(values):
        cell = ws.cell(row=row, column=offset + 1, value=value)
        cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(vertical=top, wrap_text=wrap)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        if border:
            cell.border = border
    return row + 1


def _section(ws: Worksheet, row: int, title: str, span: int) -> int:
    """A small-caps section rule spanning the sheet's used width."""
    for column in range(1, span + 1):
        cell = ws.cell(row=row, column=column, value=title.upper() if column == 1 else None)
        cell.font = Font(name="Calibri", size=9, bold=True, color=TEAL_DARK)
        cell.border = _BOTTOM_RULE
    ws.row_dimensions[row].height = 20
    return row + 1


def _table_header(ws: Worksheet, row: int, headers: list[str]) -> int:
    return _write(
        ws,
        row,
        headers,
        bold=True,
        size=10,
        color=INK_SOFT,
        fill=TEAL_WASH,
        border=_HEADER_BORDER,
    )


def _widths(ws: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _fmt(value: datetime | None) -> str:
    return value.strftime("%b %d, %Y at %H:%M") if value else "-"


def _duration(started: datetime | None, ended: datetime | None) -> str:
    if not started or not ended:
        return ""
    minutes = int(max(0, (ended - started).total_seconds()) // 60)
    if minutes < 60:
        return f"{minutes} min"
    return f"{minutes // 60}h {minutes % 60:02d}m"


def _clean(value: str | None) -> str:
    return (value or "").strip()


# Display-name substitution, passed in by the router so every sheet applies the
# same one the Insights grid does. Identity when the session has no renames.
Rename = Callable[[str], str]


def _identity(text: str) -> str:
    return text


def _speaker_lookup(participants: list[Participant]) -> dict[str, str]:
    return {p.id: p.label for p in participants}


def _attributed(question, lookup: dict[str, str]) -> str:
    raw = getattr(question, "speaker_id", None)
    return lookup.get(str(raw), "") if raw else ""


def build_summary_sheet(
    ws: Worksheet,
    *,
    session,
    questions: list,
    participants: list[Participant],
    rename: Rename = _identity,
) -> None:
    """The sheet the workbook opens on.

    Three questions in order: what came out of this call, who was in it, and
    what is still open. ``questions`` must already exclude dismissed rows -
    a dismissed insight is one the user rejected, and counting it here would
    inflate every number on the page.
    """
    ws.sheet_view.showGridLines = False
    _widths(ws, [58, 14, 34, 26])

    row = _summary_masthead(ws, session, questions, participants, rename)
    row = _at_a_glance_block(ws, row, questions)
    row = _participants_block(ws, row, questions, participants)
    _needs_attention_block(ws, row, questions, participants, rename)

    # Below the masthead, so the session name stays on screen while scrolling.
    ws.freeze_panes = "A5"


def _summary_masthead(ws, session, questions: list, participants: list, rename: Rename) -> int:
    row = _write(ws, 1, ["BACKCHANNEL BRIEFING"], bold=True, size=9, color=TEAL_DARK)
    name = rename(_clean(getattr(session, "name", ""))) or "Session"
    row = _write(ws, row, [name], bold=True, size=20)
    ws.row_dimensions[row - 1].height = 30

    started = getattr(session, "started_at", None)
    ended = getattr(session, "ended_at", None)
    meta = [f"Started {_fmt(started)}"]
    if span := _duration(started, ended):
        meta.append(span)
    meta.append(f"{len(participants)} participant{'s' if len(participants) != 1 else ''}")
    meta.append(f"{len(questions)} insight{'s' if len(questions) != 1 else ''}")
    return _write(ws, row, [" | ".join(meta)], size=10, color=INK_FAINT) + 1


def _insight_mix(questions: list) -> tuple[list[str], dict[str, int], dict[str, int]]:
    """Counts and starred counts per type, in the reading order above."""
    counts: dict[str, int] = {}
    starred: dict[str, int] = {}
    for q in questions:
        slug = getattr(q, "item_type", None) or "question"
        counts[slug] = counts.get(slug, 0) + 1
        if getattr(q, "starred", False):
            starred[slug] = starred.get(slug, 0) + 1
    ordered = sorted(
        counts,
        key=lambda slug: (
            _TYPE_ORDER.index(slug) if slug in _TYPE_ORDER else len(_TYPE_ORDER),
            slug,
        ),
    )
    return ordered, counts, starred


def _at_a_glance_block(ws, row: int, questions: list) -> int:
    row = _section(ws, row, "At a glance", 4)
    row = _table_header(ws, row, ["Insight type", "Count", "Share", "Starred"])
    ordered, counts, starred = _insight_mix(questions)
    if not ordered:
        return _write(ws, row, ["No insights captured."], size=10, color=INK_FAINT) + 1

    total = sum(counts.values())
    first = row
    for slug in ordered:
        row = _write(
            ws,
            row,
            [type_label(slug), counts[slug], f"{counts[slug] / total:.0%}", starred.get(slug, 0) or ""],
            border=_BOTTOM_RULE,
        )
    # A native data bar rather than a drawn chart: it survives a re-sort,
    # prints, and costs the file nothing.
    ws.conditional_formatting.add(
        f"B{first}:B{row - 1}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=TEAL, showValue=True),
    )
    return row + 1


def _attributed_counts(questions: list) -> dict[str, int]:
    """Insights per speaker id.

    Keyed by id, not by label: display_name is free text and two people can
    carry the same one, which would silently pool their counts.
    """
    counts: dict[str, int] = {}
    for q in questions:
        raw = getattr(q, "speaker_id", None)
        if raw:
            counts[str(raw)] = counts.get(str(raw), 0) + 1
    return counts


def _participants_block(ws, row: int, questions: list, participants: list) -> int:
    row = _section(ws, row, "Participants", 4)
    row = _table_header(ws, row, ["Participant", "Side", "Role", "Insights attributed"])
    if not participants:
        return _write(
            ws, row, ["No speakers were enrolled for this session."], size=10, color=INK_FAINT
        ) + 1

    attributed = _attributed_counts(questions)
    for person in participants:
        row = _write(
            ws,
            row,
            [person.label, person.side, person.role, attributed.get(person.id, 0)],
            border=_BOTTOM_RULE,
        )
    return row + 1


def _needs_attention_block(ws, row: int, questions: list, participants: list, rename: Rename) -> int:
    lookup = _speaker_lookup(participants)
    row = _section(ws, row, "Needs attention", 4)
    row = _table_header(ws, row, ["Item", "Type", "Why it is here", "Raised by"])
    flagged = _needs_attention(questions)
    if not flagged:
        return _write(
            ws, row, ["Nothing starred, unanswered, or contested."], size=10, color=INK_FAINT
        )

    for q, reason in flagged:
        row = _write(
            ws,
            row,
            [
                rename(_clean(q.question)),
                type_label(getattr(q, "item_type", None)),
                reason,
                _attributed(q, lookup),
            ],
            wrap=True,
            border=_BOTTOM_RULE,
        )
        ws.row_dimensions[row - 1].height = 30
    return row


def _needs_attention(questions: list) -> list[tuple[object, str]]:
    """Starred items, open questions and objections, each said once.

    Ordered by how much a reader has to do about it, not by insight type: a
    starred item was flagged by hand, an objection is live resistance, an open
    question is a gap. An item that is several of those is listed under the
    first that applies rather than repeated.
    """
    flagged: list[tuple[object, str]] = []
    for q in questions:
        if getattr(q, "starred", False):
            flagged.append((q, "Starred"))
    seen = {id(q) for q, _ in flagged}
    for q in questions:
        if id(q) in seen or (getattr(q, "item_type", None) or "") != "objection":
            continue
        flagged.append((q, "Objection raised"))
        seen.add(id(q))
    for q in questions:
        if id(q) in seen:
            continue
        if (getattr(q, "item_type", None) or "") == "question" and not getattr(q, "answered", False):
            flagged.append((q, "Unanswered"))
            seen.add(id(q))
    return flagged


def build_action_items_sheet(
    ws: Worksheet,
    *,
    questions: list,
    participants: list[Participant],
    rename: Rename = _identity,
) -> None:
    """The action items, shaped so a person can work down the column.

    There is no owner field anywhere in the schema, so this does not invent
    one. "Raised by" is the attributed speaker, which is the only ownership
    signal the data actually carries.
    """
    ws.sheet_view.showGridLines = False
    _widths(ws, [62, 20, 12, 40, 44, 18])
    lookup = _speaker_lookup(participants)
    text = lambda value: rename(_clean(value))  # noqa: E731
    items = [q for q in questions if (getattr(q, "item_type", None) or "") == "action_item"]

    row = _write(ws, 1, ["ACTION ITEMS"], bold=True, size=9, color=TEAL_DARK)
    row = _write(ws, row, [f"{len(items)} captured"], size=10, color=INK_FAINT)
    row += 1
    header_row = row
    row = _table_header(ws, row, ["Action", "Raised by", "Status", "Follow-up", "Context", "Captured"])
    for q in items:
        status = "Done" if getattr(q, "answered", False) else "Open"
        row = _write(
            ws,
            row,
            [
                text(q.question),
                _attributed(q, lookup),
                status,
                text(getattr(q, "followup_question", "")) or text(getattr(q, "answer_summary", "")),
                text(getattr(q, "source_context", "")),
                _created(q),
            ],
            wrap=True,
            border=_BOTTOM_RULE,
        )
        ws.row_dimensions[row - 1].height = 32
        if getattr(q, "starred", False):
            ws.cell(row=row - 1, column=1).font = Font(name="Calibri", size=11, bold=True, color=INK)
    if not items:
        _write(ws, row, ["No action items were captured on this call."], size=10, color=INK_FAINT)
    ws.freeze_panes = f"A{header_row + 1}"


def build_opportunities_sheet(
    ws: Worksheet,
    *,
    questions: list,
    participants: list[Participant],
    rename: Rename = _identity,
) -> None:
    """Each opportunity beside what the specialist matched it to."""
    ws.sheet_view.showGridLines = False
    _widths(ws, [56, 56, 40, 20, 18])
    lookup = _speaker_lookup(participants)
    text = lambda value: rename(_clean(value))  # noqa: E731
    items = [q for q in questions if (getattr(q, "item_type", None) or "") == "opportunity"]
    matched = sum(1 for q in items if _clean(getattr(q, "offering_match", "")))  # rename cannot empty a value

    row = _write(ws, 1, ["OPPORTUNITIES"], bold=True, size=9, color=TEAL_DARK)
    row = _write(
        ws,
        row,
        [f"{len(items)} surfaced | {matched} matched to an offering"],
        size=10,
        color=INK_FAINT,
    )
    row += 1
    header_row = row
    row = _table_header(ws, row, ["Opportunity", "Offering match", "Why it matters", "Raised by", "Captured"])
    for q in items:
        match = text(getattr(q, "offering_match", ""))
        row = _write(
            ws,
            row,
            [
                text(q.question),
                match or "Not matched",
                text(getattr(q, "rationale", "")),
                _attributed(q, lookup),
                _created(q),
            ],
            wrap=True,
            border=_BOTTOM_RULE,
        )
        ws.row_dimensions[row - 1].height = 34
        if not match:
            ws.cell(row=row - 1, column=2).font = Font(name="Calibri", size=11, color=INK_FAINT, italic=True)
        if notes := text(getattr(q, "enrichment_notes", "")):
            row = _write(ws, row, ["", notes], size=10, color=INK_SOFT, wrap=True, border=_BOTTOM_RULE)
    if not items:
        _write(ws, row, ["No opportunities were surfaced on this call."], size=10, color=INK_FAINT)
    ws.freeze_panes = f"A{header_row + 1}"


def _created(question) -> str:
    stamp = getattr(question, "created_at", None)
    return stamp.strftime("%Y-%m-%d %H:%M") if stamp else ""
