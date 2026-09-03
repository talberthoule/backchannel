import os
import tempfile
import uuid

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import Directive, Question, Session, Speaker, TranscriptEntry
from app.services import briefing_document, insight_workbook, runtime_activity
from app.services.pii import shield, vault
from app.services.briefing_synthesis import get_session_synthesis
from sqlalchemy import select

router = APIRouter(prefix="/api/sessions/{session_id}/artifacts", tags=["artifacts"])


async def _stream_and_cleanup(file_path: str, media_type: str, filename: str):
    """Stream a temp file to the client, then delete it immediately after."""
    with runtime_activity.track("artifact export"):
        try:
            with open(file_path, "rb") as f:
                while chunk := f.read(8192):
                    yield chunk
        finally:
            try:
                os.unlink(file_path)
            except OSError:
                pass


_INSIGHT_TEXT_FIELDS = (
    "question", "rationale", "source_context", "answer_summary",
    "followup_question", "enrichment_notes", "offering_match", "lens_label",
)


def _stream_bytes(data: bytes):
    """Stream in-memory bytes — nothing ever touches disk."""
    with runtime_activity.track("artifact export"):
        yield data


@router.get("/transcript-export")
async def export_transcript(session_id: uuid.UUID, reveal: bool = False, db: AsyncSession = Depends(get_db)):
    """Export session transcript as a downloadable text file. Generated in-memory, never stored."""
    session = await db.get(Session, session_id)
    if not session:
        raise HTTPException(404, "Session not found")

    result = await db.execute(
        select(TranscriptEntry)
        .where(TranscriptEntry.session_id == session_id)
        .order_by(TranscriptEntry.sequence)
    )
    entries = result.scalars().all()

    # Load speakers for display name substitution
    spk_result = await db.execute(
        select(Speaker).where(Speaker.session_id == session_id)
    )
    speakers = spk_result.scalars().all()
    speaker_name_map = {}
    for s in speakers:
        if s.display_name and s.display_name_enabled:
            speaker_name_map[s.name] = s.display_name

    def _apply_names(text: str) -> str:
        for original, display in speaker_name_map.items():
            text = text.replace(original, display)
        return text

    lines = [f"Transcript: {session.name}", f"Date: {session.started_at or session.created_at}", ""]
    for entry in entries:
        ts = entry.timestamp.strftime("%H:%M:%S") if entry.timestamp else ""
        lines.append(f"[{ts}] {_apply_names(entry.text)}")

    text = "\n".join(lines)
    if reveal:
        text = await shield.reveal_text(db, session_id, text, route="transcript-export")
    content = text.encode("utf-8")
    filename = f"transcript-{session.name.replace(' ', '_')}.txt"

    return StreamingResponse(
        _stream_bytes(content),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/questions-export")
async def export_insights(
    session_id: uuid.UUID,
    reveal: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """Export all session insights."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.filters import AutoFilter, CustomFilter, CustomFilters, FilterColumn
    from openpyxl.utils import get_column_letter

    session = await db.get(Session, session_id)
    if not session:
        raise HTTPException(404, "Session not found")

    result = await db.execute(
        select(Question)
        .where(Question.session_id == session_id)
        .order_by(Question.vote.desc(), Question.created_at)
    )
    questions = result.scalars().all()

    wb = Workbook()
    # Three reader-facing sheets come first so the workbook opens on a summary
    # rather than on an eighteen-column grid (ALP-369); Insights is unchanged.
    summary_ws = wb.active
    summary_ws.title = "Summary"
    actions_ws = wb.create_sheet("Action Items")
    opportunities_ws = wb.create_sheet("Opportunities")
    ws = wb.create_sheet("Insights")

    # Load speakers for source_context name replacement
    spk_result = await db.execute(
        select(Speaker).where(Speaker.session_id == session_id).order_by(Speaker.created_at)
    )
    speakers = spk_result.scalars().all()
    speaker_name_map = {}
    for s in speakers:
        if s.display_name and s.display_name_enabled:
            speaker_name_map[s.name] = s.display_name

    def _apply_speaker_names(text: str) -> str:
        for original, display in speaker_name_map.items():
            text = text.replace(original, display)
        return text

    # A plain view of the roster, built before any reveal so the ORM rows are
    # never mutated. Names may be vault tokens; they are revealed below only
    # when the caller asked, exactly like the insight text.
    participants = [
        insight_workbook.Participant(
            id=str(s.id),
            label=(s.display_name if s.display_name and s.display_name_enabled else s.name) or "Unknown",
            role=s.role or "",
            side="Team" if s.speaker_type == "team" else "External",
        )
        for s in speakers
    ]

    if reveal:
        # Detached first: a revealed value must never be flushed back into
        # the row it came from.
        mapping = await vault.reveal_map(db, session_id)
        revealed_total = 0
        for q in questions:
            db.expunge(q)
            for attr in _INSIGHT_TEXT_FIELDS:
                value = getattr(q, attr, None)
                if isinstance(value, str) and value:
                    value, count = shield.substitute(value, mapping)
                    revealed_total += count
                    setattr(q, attr, value)
        # The roster is a copy, so this reveals the labels the summary sheet
        # prints without touching a Speaker row.
        revealed_participants = []
        for person in participants:
            label, count = shield.substitute(person.label, mapping)
            revealed_total += count
            revealed_participants.append(
                insight_workbook.Participant(id=person.id, label=label, role=person.role, side=person.side)
            )
        participants = revealed_participants
        if revealed_total:
            await shield.record_reveal(session_id, "questions-export", revealed_total)

    headers = [
        "Type",
        "Lens",
        "Insight",
        "Rationale",
        "Source Context",
        "Vote",
        "Starred",
        "Answered",
        "Answer Summary",
        "Needs Follow-up",
        "Follow-up Question",
        "Offering Match",
        "Enhanced",
        "Dismissed",
        "Enrichment Notes",
        "Agent Source",
        "Revision Count",
        "Created At",
    ]
    ws.append(headers)

    type_labels = {
        "question": "Question",
        "observation": "Observation",
        "opportunity": "Opportunity",
        "action_item": "Action Item",
        "objection": "Objection",
    }

    for q in questions:
        ws.append([
            type_labels.get(q.item_type, (q.item_type or "question").replace("_", " ").title()),
            q.lens_label or "",
            _apply_speaker_names(q.question),
            _apply_speaker_names(q.rationale),
            _apply_speaker_names(q.source_context),
            q.vote or 0,
            q.starred,
            q.answered,
            _apply_speaker_names(q.answer_summary) if q.answer_summary else "",
            q.needs_followup,
            q.followup_question or "",
            q.offering_match or "",
            q.enhanced,
            q.dismissed,
            q.enrichment_notes or "",
            q.agent_source or "",
            q.revision_count,
            q.created_at.strftime("%Y-%m-%d %H:%M:%S") if q.created_at else "",
        ])

    # Format as an Excel Table
    if questions:
        last_col = get_column_letter(len(headers))
        last_row = len(questions) + 1
        table_ref = f"A1:{last_col}{last_row}"
        table = Table(displayName="Insights", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # Default filter: Dismissed column (index 11) shows only FALSE
        dismissed_col_idx = headers.index("Dismissed")
        dismissed_filter = FilterColumn(
            colId=dismissed_col_idx,
            customFilters=CustomFilters(customFilter=[
                CustomFilter(val="FALSE"),  # openpyxl custom filter matches string repr
            ]),
        )
        table.autoFilter = AutoFilter(ref=table_ref)
        table.autoFilter.filterColumn.append(dismissed_filter)

        # Hide dismissed rows so the filter appears pre-applied when opened
        for row_idx in range(2, last_row + 1):
            dismissed_cell = ws.cell(row=row_idx, column=dismissed_col_idx + 1)
            if dismissed_cell.value is True:
                ws.row_dimensions[row_idx].hidden = True

        ws.add_table(table)

    # Auto-size columns (approximate)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # The reader-facing sheets are built from the same rows the grid holds, and
    # from the same already-revealed (or still tokenized) strings, so they can
    # never show a value the Insights sheet would not. Dismissed insights are
    # left out: the grid hides them behind a filter, and counting a rejected
    # item would inflate every number on the summary.
    live = [q for q in questions if not q.dismissed]
    insight_workbook.build_summary_sheet(
        summary_ws,
        session=session,
        questions=live,
        participants=participants,
        rename=_apply_speaker_names,
    )
    insight_workbook.build_action_items_sheet(
        actions_ws, questions=live, participants=participants, rename=_apply_speaker_names
    )
    insight_workbook.build_opportunities_sheet(
        opportunities_ws, questions=live, participants=participants, rename=_apply_speaker_names
    )

    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    filename = f"insights-{session.name.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        _stream_bytes(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/summary-export")
async def export_summary(session_id: uuid.UUID, reveal: bool = False, db: AsyncSession = Depends(get_db)):
    """Export the session briefing as one self-contained HTML file.

    Generated in memory, never stored. The markup lives in
    ``services/briefing_document``; this handler owns the queries and the
    reveal rule, so a renderer can only draw what was loaded for it.
    """
    session = await db.get(Session, session_id)
    if not session:
        raise HTTPException(404, "Session not found")

    synthesis = await get_session_synthesis(session_id, mode="post_call")
    data = await _briefing_data(session, session_id, db, synthesis)

    if briefing_document.synthesis_is_usable(synthesis):
        html = briefing_document.render_briefing(data)
        prefix = "briefing"
    else:
        data.synthesis = None
        html = briefing_document.render_record(data)
        prefix = "summary"

    if reveal:
        # The whole rendered document is revealed in one pass. Every vault
        # token reaches the page as literal text -- a token carries no
        # HTML-special characters, so escaping leaves it intact -- and without
        # this flag nothing here has ever seen a value to leak.
        html = await shield.reveal_text(db, session_id, html, route="summary-export")
    content = html.encode("utf-8")
    filename = f"{prefix}-{session.name.replace(' ', '_')}.html"

    return StreamingResponse(
        _stream_bytes(content),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _briefing_data(
    session,
    session_id: uuid.UUID,
    db: AsyncSession,
    synthesis,
) -> briefing_document.BriefingData:
    """Load the briefing's supporting rows.

    Both renderers get the same set. The synthesized briefing used to draw
    from the synthesis alone, which is why it could show a confidence level
    but not who was in the room or what any claim rested on (ALP-370).
    Speakers are ordered by creation so a roster reads the same way it does
    everywhere else in the app.
    """
    questions = (
        await db.execute(
            select(Question)
            .where(Question.session_id == session_id, Question.dismissed.is_(False))
            .order_by(Question.vote.desc(), Question.created_at)
        )
    ).scalars().all()
    speakers = (
        await db.execute(
            select(Speaker).where(Speaker.session_id == session_id).order_by(Speaker.created_at)
        )
    ).scalars().all()
    transcript = (
        await db.execute(
            select(TranscriptEntry)
            .where(TranscriptEntry.session_id == session_id)
            .order_by(TranscriptEntry.sequence)
        )
    ).scalars().all()
    directives = (
        await db.execute(
            select(Directive)
            .where(Directive.session_id == session_id, Directive.active.is_(True))
            .order_by(Directive.created_at)
        )
    ).scalars().all()
    return briefing_document.BriefingData(
        session=session,
        synthesis=synthesis,
        questions=list(questions),
        speakers=list(speakers),
        transcript=list(transcript),
        directives=list(directives),
    )
