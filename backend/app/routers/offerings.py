"""Offerings CRUD + CSV/Excel import endpoint."""

import csv
import io
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import Offering
from app.schemas import OfferingCreate, OfferingOut, OfferingUpdate

router = APIRouter(prefix="/api/offerings", tags=["offerings"])


@router.get("", response_model=list[OfferingOut])
async def list_offerings(
    vendor: str | None = None,
    category: str | None = None,
    tag: str | None = None,
    active_only: bool = True,
    db: AsyncSession = Depends(get_db),
):
    query = select(Offering)
    if active_only:
        query = query.where(Offering.active.is_(True))
    if vendor:
        query = query.where(Offering.vendor == vendor)
    if category:
        query = query.where(Offering.category == category)
    if tag:
        query = query.where(Offering.tags.ilike(f"%{tag}%"))
    query = query.order_by(Offering.vendor, Offering.category, Offering.product_name)
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/vendors", response_model=list[str])
async def list_vendors(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Offering.vendor).distinct().order_by(Offering.vendor)
    )
    return [r[0] for r in result.all()]


@router.get("/categories", response_model=list[str])
async def list_categories(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Offering.category).distinct().order_by(Offering.category)
    )
    return [r[0] for r in result.all()]


@router.get("/tags", response_model=list[str])
async def list_tags(db: AsyncSession = Depends(get_db)):
    """Return the distinct tags currently in use across the catalog."""
    result = await db.execute(select(Offering.tags))
    tags: set[str] = set()
    for (value,) in result.all():
        for tag in (value or "").split(","):
            tag = tag.strip()
            if tag:
                tags.add(tag)
    return sorted(tags)


@router.post("", response_model=OfferingOut)
async def create_offering(body: OfferingCreate, db: AsyncSession = Depends(get_db)):
    offering = Offering(**body.model_dump())
    db.add(offering)
    await db.commit()
    await db.refresh(offering)
    return offering


@router.patch("/{offering_id}", response_model=OfferingOut)
async def update_offering(
    offering_id: uuid.UUID, body: OfferingUpdate, db: AsyncSession = Depends(get_db)
):
    offering = await db.get(Offering, offering_id)
    if not offering:
        raise HTTPException(404, "Offering not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(offering, field, value)
    offering.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(offering)
    return offering


@router.delete("/{offering_id}")
async def delete_offering(offering_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    offering = await db.get(Offering, offering_id)
    if not offering:
        raise HTTPException(404, "Offering not found")
    await db.delete(offering)
    await db.commit()
    return {"ok": True}


@router.post("/import")
async def import_offerings(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """Import offerings from CSV or Excel file.

    Expected columns: vendor, product_name, category, subcategory, description, use_cases, note, tags.
    Legacy column names (discipline, delivery_model, practice) are accepted as aliases.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(400, "openpyxl is required for Excel imports. Install it with: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val) if val is not None else ""
            rows.append(row_dict)
    else:
        raise HTTPException(400, "Unsupported file format. Use .csv or .xlsx")

    created = 0
    skipped = 0
    for row in rows:
        normalized = _normalize_offering_import_row(row)
        vendor = normalized["vendor"]
        product_name = normalized["product_name"]
        if not vendor or not product_name:
            skipped += 1
            continue

        # Check for duplicate
        existing = await db.execute(
            select(Offering).where(
                Offering.vendor == vendor,
                Offering.product_name == product_name,
            )
        )
        if existing.scalar_one_or_none():
            skipped += 1
            continue

        offering = Offering(**normalized)
        db.add(offering)
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "total_rows": len(rows)}


def _normalize_offering_import_row(row: dict) -> dict[str, str]:
    return {
        "vendor": _field(row, "vendor") or "In-House",
        "product_name": _field(row, "product_name"),
        "category": _field(row, "category"),
        # Legacy files carried a separate "discipline" column; fold it into subcategory.
        "subcategory": _field(row, "subcategory") or _field(row, "discipline"),
        "description": _field(row, "description"),
        "use_cases": _field(row, "use_cases"),
        "note": _field(row, "note") or _field(row, "delivery_model"),
        "tags": _field(row, "tags") or _field(row, "practice"),
    }


def _field(row: dict, key: str) -> str:
    return str(row.get(key, "") or "").strip()


@router.post("/seed")
async def seed_offerings(replace: bool = False, db: AsyncSession = Depends(get_db)):
    """Seed the database with sample offerings.

    Pass ?replace=true to delete all existing offerings and reseed.
    """
    count = await db.scalar(select(func.count(Offering.id)))
    if count and count > 0 and not replace:
        return {"message": f"Database already has {count} offerings. Pass ?replace=true to reseed."}

    if replace and count:
        from sqlalchemy import delete
        await db.execute(delete(Offering))

    offerings = _get_seed_data()
    for o in offerings:
        db.add(Offering(**o))
    await db.commit()
    return {"created": len(offerings), "replaced": replace}


def _get_seed_data() -> list[dict]:
    """Fictional sample catalog: well-known vendors, generic tags and notes."""
    return [
        # -- Microsoft --
        {"vendor": "Microsoft", "product_name": "Microsoft 365 E5", "category": "Productivity", "subcategory": "Suite Licensing", "description": "Productivity suite with advanced security, compliance, and analytics add-ons.", "use_cases": "Collaboration, email security, insider risk, eDiscovery", "note": "Strong fit when a customer wants to consolidate point security tools.", "tags": "Hybrid Work, Compliance"},
        {"vendor": "Microsoft", "product_name": "Microsoft Entra ID", "category": "Security", "subcategory": "Identity & Access", "description": "Cloud identity and access management with SSO, MFA, and conditional access.", "use_cases": "Single sign-on, MFA rollout, conditional access, identity governance", "note": "Pairs well with a Zero Trust assessment engagement.", "tags": "Zero Trust, Compliance"},
        {"vendor": "Microsoft", "product_name": "Microsoft Sentinel", "category": "Security", "subcategory": "SIEM", "description": "Cloud-native SIEM and SOAR with built-in analytics and threat intelligence.", "use_cases": "Log analytics, threat detection, SOC automation, compliance reporting", "note": "Often attached to our managed detection service.", "tags": "Security Operations, Cloud"},
        {"vendor": "Microsoft", "product_name": "Azure Landing Zone Deployment", "category": "Professional Services", "subcategory": "Cloud Foundations", "description": "Design and deployment of a governed Azure environment ready for workload migration.", "use_cases": "Cloud adoption, governance guardrails, migration readiness", "note": "Fixed-scope engagement, typically four to six weeks.", "tags": "Cloud, Consulting"},

        # -- AWS --
        {"vendor": "AWS", "product_name": "AWS Migration Program", "category": "Cloud", "subcategory": "Migration", "description": "Assessment, planning, and execution of workload migrations to AWS.", "use_cases": "Data center exit, hardware refresh avoidance, app modernization", "note": "Migration funding credits may offset services cost.", "tags": "Cloud, Consulting"},
        {"vendor": "AWS", "product_name": "Amazon Connect", "category": "Collaboration", "subcategory": "Contact Center", "description": "Cloud contact center with intelligent routing, IVR, and real-time analytics.", "use_cases": "Contact center modernization, seasonal scaling, agent analytics", "note": "Usage-based pricing suits variable call volumes.", "tags": "Customer Experience, Cloud"},
        {"vendor": "AWS", "product_name": "AWS Cost Optimization Review", "category": "Professional Services", "subcategory": "FinOps", "description": "Structured review of AWS spend with a prioritized savings roadmap.", "use_cases": "Budget overruns, reserved instance planning, rightsizing", "note": "Two-week assessment with an executive readout.", "tags": "Cloud, Cost Optimization, Consulting"},

        # -- Cisco --
        {"vendor": "Cisco", "product_name": "Cisco Meraki", "category": "Networking", "subcategory": "Cloud-Managed Networking", "description": "Cloud-managed switches, wireless, security appliances, and cameras.", "use_cases": "Multi-site rollouts, simplified management, guest Wi-Fi", "note": "Licensing renews annually; co-termination available.", "tags": "Networking, Managed Service"},
        {"vendor": "Cisco", "product_name": "Cisco Catalyst Center", "category": "Networking", "subcategory": "Network Automation", "description": "Network controller for automation, assurance, and policy across campus networks.", "use_cases": "Network automation, configuration drift, assurance analytics", "note": "Requires current Catalyst hardware; check install base first.", "tags": "Networking, Automation"},

        # -- Fortinet --
        {"vendor": "Fortinet", "product_name": "FortiGate NGFW", "category": "Security", "subcategory": "Network Security", "description": "Next-generation firewall with IPS, web filtering, and VPN.", "use_cases": "Perimeter security, branch security, VPN consolidation", "note": "Available as hardware or virtual appliance.", "tags": "Networking, Data Protection"},
        {"vendor": "Fortinet", "product_name": "Fortinet Secure SD-WAN", "category": "Networking", "subcategory": "SD-WAN", "description": "Integrated SD-WAN and security on a single platform.", "use_cases": "Branch connectivity, MPLS cost reduction, application steering", "note": "Strong option when firewall refresh and WAN projects align.", "tags": "Networking, Cost Optimization"},

        # -- Okta --
        {"vendor": "Okta", "product_name": "Okta Workforce Identity", "category": "Security", "subcategory": "Identity & Access", "description": "Independent identity platform with SSO, MFA, and lifecycle management.", "use_cases": "App consolidation, contractor access, identity lifecycle automation", "note": "Common coexistence questions with Microsoft Entra; position by app estate.", "tags": "Zero Trust, Hybrid Work"},

        # -- Datadog --
        {"vendor": "Datadog", "product_name": "Datadog Observability Platform", "category": "Observability", "subcategory": "Monitoring & APM", "description": "Unified metrics, traces, and logs with dashboards and alerting.", "use_cases": "Cloud migration visibility, incident response, SLO tracking", "note": "Cost scales with hosts and log volume; set ingestion budgets early.", "tags": "Observability, Cloud"},

        # -- Zoom --
        {"vendor": "Zoom", "product_name": "Zoom Workplace", "category": "Collaboration", "subcategory": "Meetings & Chat", "description": "Video meetings, team chat, whiteboard, and phone in one client.", "use_cases": "Hybrid meetings, phone system replacement, room modernization", "note": "Competitive displacement pricing often available.", "tags": "Hybrid Work, Customer Experience"},

        # -- Veeam --
        {"vendor": "Veeam", "product_name": "Veeam Data Platform", "category": "Data Management", "subcategory": "Backup & Recovery", "description": "Backup, recovery, and ransomware resilience for hybrid environments.", "use_cases": "Ransomware recovery, immutable backups, cloud tiering", "note": "Recovery-time stories resonate well with executive audiences.", "tags": "Data Protection, Compliance"},

        # -- In-House services --
        {"vendor": "In-House", "product_name": "Managed Detection & Response", "category": "Managed Services", "subcategory": "Security Operations", "description": "24/7 monitoring, triage, and guided response across endpoints and cloud.", "use_cases": "SOC coverage gaps, alert fatigue, compliance evidence", "note": "Thirty-day onboarding; works with the customer's existing tooling.", "tags": "Security Operations, Managed Service"},
        {"vendor": "In-House", "product_name": "Managed Cloud Operations", "category": "Managed Services", "subcategory": "Cloud Operations", "description": "Ongoing management of cloud infrastructure: patching, backup, and cost hygiene.", "use_cases": "Staff shortages, patching discipline, monthly cost reviews", "note": "Tiered by environment size; includes quarterly optimization reviews.", "tags": "Cloud, Managed Service, Cost Optimization"},
        {"vendor": "In-House", "product_name": "Help Desk as a Service", "category": "Managed Services", "subcategory": "End-User Support", "description": "White-labeled tier 1-2 support desk with SLA-backed response times.", "use_cases": "After-hours coverage, ticket backlog, onboarding surges", "note": "Per-seat pricing; ninety-day minimum term.", "tags": "Managed Service, Hybrid Work"},
        {"vendor": "In-House", "product_name": "Security Program Assessment", "category": "Professional Services", "subcategory": "Advisory", "description": "Point-in-time review of security posture with a prioritized remediation roadmap.", "use_cases": "Board reporting, cyber insurance requirements, audit preparation", "note": "Findings mapped to NIST CSF; executive readout included.", "tags": "Consulting, Compliance"},
        {"vendor": "In-House", "product_name": "Network Assessment & Design", "category": "Professional Services", "subcategory": "Advisory", "description": "Discovery and design engagement covering LAN, WAN, and wireless.", "use_cases": "Office moves, refresh planning, performance complaints", "note": "Deliverables feed directly into an implementation statement of work.", "tags": "Consulting, Networking"},
        {"vendor": "In-House", "product_name": "AI Readiness Workshop", "category": "Professional Services", "subcategory": "Advisory", "description": "Half-day workshop identifying practical AI use cases and data prerequisites.", "use_cases": "Executive AI strategy, pilot selection, data readiness", "note": "Popular door-opener; frequently leads to a pilot engagement.", "tags": "Consulting, Automation"},
    ]
