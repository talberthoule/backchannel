"""Knowledge source and record CRUD + import/upload endpoints."""

import csv
import io
import json
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import AgentConfig, KnowledgeRecord, KnowledgeSource
from app.schemas import (
    KnowledgeRecordCreate,
    KnowledgeRecordOut,
    KnowledgeRecordUpdate,
    KnowledgeSourceCreate,
    KnowledgeSourceOut,
    KnowledgeSourceUpdate,
)
from app.services.knowledge import USER_SOURCE_TYPES
from app.services.markdown_conversion import (
    MarkdownConversionError,
    convert_to_markdown,
    file_extension,
)

router = APIRouter(prefix="/api/knowledge", tags=["knowledge"])


async def _get_source_or_404(source_id: uuid.UUID, db: AsyncSession) -> KnowledgeSource:
    source = await db.get(KnowledgeSource, source_id)
    if not source:
        raise HTTPException(404, "Knowledge source not found")
    return source


def _require_user_source(source: KnowledgeSource):
    if source.source_type not in USER_SOURCE_TYPES:
        raise HTTPException(400, "Records for this source are managed elsewhere (see the Offerings manager)")


async def _source_out(source: KnowledgeSource, db: AsyncSession) -> KnowledgeSourceOut:
    count = await db.scalar(
        select(func.count()).select_from(KnowledgeRecord).where(KnowledgeRecord.source_id == source.id)
    )
    out = KnowledgeSourceOut.model_validate(source)
    out.record_count = int(count or 0)
    return out


@router.get("", response_model=list[KnowledgeSourceOut])
async def list_sources(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(KnowledgeSource, func.count(KnowledgeRecord.id))
        .outerjoin(KnowledgeRecord, KnowledgeRecord.source_id == KnowledgeSource.id)
        .group_by(KnowledgeSource.id)
        .order_by(KnowledgeSource.created_at)
    )
    out = []
    for source, count in result.all():
        item = KnowledgeSourceOut.model_validate(source)
        item.record_count = int(count or 0)
        out.append(item)
    return out


@router.post("", response_model=KnowledgeSourceOut)
async def create_source(body: KnowledgeSourceCreate, db: AsyncSession = Depends(get_db)):
    if body.source_type not in USER_SOURCE_TYPES:
        raise HTTPException(400, f"source_type must be one of {sorted(USER_SOURCE_TYPES)}")
    source = KnowledgeSource(**body.model_dump())
    db.add(source)
    await db.commit()
    await db.refresh(source)
    return await _source_out(source, db)


@router.get("/{source_id}", response_model=KnowledgeSourceOut)
async def get_source(source_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    source = await _get_source_or_404(source_id, db)
    return await _source_out(source, db)


@router.patch("/{source_id}", response_model=KnowledgeSourceOut)
async def update_source(
    source_id: uuid.UUID, body: KnowledgeSourceUpdate, db: AsyncSession = Depends(get_db)
):
    source = await _get_source_or_404(source_id, db)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(source, field, value)
    source.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(source)
    return await _source_out(source, db)


@router.delete("/{source_id}")
async def delete_source(source_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    source = await _get_source_or_404(source_id, db)
    if source.source_type == "offerings":
        raise HTTPException(400, "The built-in offerings source cannot be deleted")
    result = await db.execute(
        select(AgentConfig.slug).where(AgentConfig.knowledge_source_ids.contains(str(source_id)))
    )
    referencing = [row[0] for row in result.all()]
    if referencing:
        raise HTTPException(
            400,
            f"Source is in use by agent(s): {', '.join(referencing)}. Point them elsewhere first.",
        )
    await db.execute(delete(KnowledgeRecord).where(KnowledgeRecord.source_id == source_id))
    await db.delete(source)
    await db.commit()
    return {"ok": True}


# --- Records ---

@router.get("/{source_id}/records", response_model=list[KnowledgeRecordOut])
async def list_records(source_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    await _get_source_or_404(source_id, db)
    result = await db.execute(
        select(KnowledgeRecord)
        .where(KnowledgeRecord.source_id == source_id)
        .order_by(KnowledgeRecord.title, KnowledgeRecord.created_at)
    )
    return result.scalars().all()


@router.post("/{source_id}/records", response_model=KnowledgeRecordOut)
async def create_record(
    source_id: uuid.UUID, body: KnowledgeRecordCreate, db: AsyncSession = Depends(get_db)
):
    source = await _get_source_or_404(source_id, db)
    _require_user_source(source)
    record = KnowledgeRecord(source_id=source_id, **body.model_dump())
    db.add(record)
    await db.commit()
    await db.refresh(record)
    return record


@router.patch("/records/{record_id}", response_model=KnowledgeRecordOut)
async def update_record(
    record_id: uuid.UUID, body: KnowledgeRecordUpdate, db: AsyncSession = Depends(get_db)
):
    record = await db.get(KnowledgeRecord, record_id)
    if not record:
        raise HTTPException(404, "Knowledge record not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(record, field, value)
    record.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(record)
    return record


@router.delete("/records/{record_id}")
async def delete_record(record_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    record = await db.get(KnowledgeRecord, record_id)
    if not record:
        raise HTTPException(404, "Knowledge record not found")
    await db.delete(record)
    await db.commit()
    return {"ok": True}


# --- Import / upload ---

def normalize_record_import_row(row: dict) -> dict:
    """Map an import row to record fields; extra columns go into meta JSON."""
    normalized = {str(k or "").strip().lower().replace(" ", "_"): str(v or "").strip() for k, v in row.items()}
    title = normalized.pop("title", "")
    body = normalized.pop("body", "")
    extras = {k: v for k, v in normalized.items() if v}
    return {"title": title, "body": body, "meta": json.dumps(extras) if extras else "{}"}


@router.post("/{source_id}/records/import")
async def import_records(
    source_id: uuid.UUID, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)
):
    """Import records from CSV or Excel. Expected columns: title, body.
    Any additional columns are stored in the record's meta JSON."""
    source = await _get_source_or_404(source_id, db)
    _require_user_source(source)

    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(400, "openpyxl is required for Excel imports. Install it with: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val) if val is not None else ""
            rows.append(row_dict)
    else:
        raise HTTPException(400, "Unsupported file format. Use .csv or .xlsx")

    created = 0
    skipped = 0
    for row in rows:
        normalized = normalize_record_import_row(row)
        if not normalized["title"] or not normalized["body"]:
            skipped += 1
            continue

        existing = await db.execute(
            select(KnowledgeRecord).where(
                KnowledgeRecord.source_id == source_id,
                KnowledgeRecord.title == normalized["title"],
            )
        )
        if existing.scalars().first():
            skipped += 1
            continue

        db.add(KnowledgeRecord(source_id=source_id, **normalized))
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "total_rows": len(rows)}


@router.post("/{source_id}/files")
async def upload_knowledge_file(
    source_id: uuid.UUID, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)
):
    """Upload a document; it is converted to Markdown (via MarkItDown) and
    stored as one record. Only the Markdown text is persisted — the original
    file is never written to the file system."""
    source = await _get_source_or_404(source_id, db)
    _require_user_source(source)

    content = await file.read()
    filename = file.filename or "upload"
    ext = file_extension(filename)

    try:
        markdown = convert_to_markdown(content, filename)
    except MarkdownConversionError as e:
        raise HTTPException(400, str(e))

    record = KnowledgeRecord(
        source_id=source_id,
        title=filename,
        body=markdown,
        meta=json.dumps({
            "filename": filename,
            "format": ext,
            "original_bytes": len(content),
            "converted_to_markdown": True,
        }),
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)
    return {"record_id": str(record.id), "title": record.title, "chars": len(record.body)}
